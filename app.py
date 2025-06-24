import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook

# Módulos locales
from db_ingresos import (
    insertar_ingreso,
    obtener_ingresos,
    eliminar_ingreso,
    actualizar_ingreso
)

from db_gastos import (
    insertar_gasto,
    obtener_gastos,
    eliminar_gasto,
    actualizar_gasto
)

from exportador_pdf import PDF


st.set_page_config(page_title="Sistema Iglesia Restauración", layout="centered")

# -------------------- MENÚ DE PESTAÑAS --------------------
menu = st.selectbox(
    "Selecciona una sección",
    [
        "📥 Registro de Ingresos",
        "💸 Registro de Gastos",
        "📊 Reporte General",
        "📄 Exportar PDF",
        
    ]
)

# -------------------- PESTAÑA: INGRESOS --------------------
if menu == "📥 Registro de Ingresos":
    st.title("📥 Registro de Ingresos")

    # ---------- FORMULARIO PARA NUEVO INGRESO ----------
    st.subheader("Agregar nuevo ingreso")
    with st.form("form_nuevo_ingreso"):
        nueva_fecha = st.date_input("Fecha")
        nuevo_concepto = st.selectbox("Concepto", ["Diezmo", "Ofrenda", "Cocina", "Otro"])
        nuevo_monto = st.number_input("Monto (₡)", min_value=0.0, step=1000.0, format="%.2f")
        nueva_observacion = st.text_area("Observación (opcional)")
        enviar = st.form_submit_button("Registrar")

        if enviar:
            resultado = insertar_ingreso(str(nueva_fecha), nuevo_concepto, nuevo_monto, nueva_observacion)
            if resultado.data:
                st.success("✅ Ingreso registrado exitosamente")
                st.rerun()
            else:
                st.error(f"❌ Error al registrar: {resultado.error}")

    # ---------- LISTADO DE INGRESOS + BOTÓN DE DESCARGA ----------
    st.subheader("📋 Ingresos registrados")
    ingresos = obtener_ingresos()

    if ingresos:
        df = pd.DataFrame(ingresos)

        # Formatear fecha y monto
        df["fecha"] = pd.to_datetime(df["fecha"]).dt.strftime("%d/%m/%Y")
        df["monto"] = df["monto"].map(lambda x: round(x, 2))

        # Generar archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Ingresos")

        st.download_button(
            label="⬇️ Descargar respaldo en Excel",
            data=output.getvalue(),
            file_name="respaldo_ingresos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Mostrar cada ingreso
        for ingreso in ingresos:
            with st.container():
                id_actual = ingreso['id']
                editando = st.session_state.get(f"edit_{id_actual}", False)

                if editando:
                    st.markdown(f"### ✏️ Editando ingreso ID {id_actual}")
                    fecha = st.date_input("Fecha", value=pd.to_datetime(ingreso["fecha"]), key=f"fecha_{id_actual}")
                    concepto = st.selectbox("Concepto", ["Diezmo", "Ofrenda", "Cocina", "Otro"],
                                            index=["Diezmo", "Ofrenda", "Cocina", "Otro"].index(ingreso["concepto"]),
                                            key=f"concepto_{id_actual}")
                    monto = st.number_input("Monto (₡)", min_value=0.0, value=float(ingreso["monto"]),
                                            format="%.2f", key=f"monto_{id_actual}")
                    observacion = st.text_input("Observación", value=ingreso["observacion"], key=f"obs_{id_actual}")
                    col1, col2 = st.columns([1, 1])
                    if col1.button("💾 Guardar", key=f"guardar_{id_actual}"):
                        actualizar_ingreso(id_actual, str(fecha), concepto, monto, observacion)
                        st.session_state[f"edit_{id_actual}"] = False
                        st.success("✅ Ingreso actualizado")
                        st.rerun()
                    if col2.button("❌ Cancelar", key=f"cancelar_{id_actual}"):
                        st.session_state[f"edit_{id_actual}"] = False
                        st.rerun()

                else:
                    fecha_formateada = pd.to_datetime(ingreso['fecha']).strftime("%d/%m/%Y")
                    cols = st.columns([1, 2, 2, 2, 3, 1, 1])
                    cols[0].markdown(f"**ID:** {ingreso['id']}")
                    cols[1].markdown(f"📅 {fecha_formateada}")
                    cols[2].markdown(f"📄 {ingreso['concepto']}")
                    cols[3].markdown(f"💰 ₡{ingreso['monto']:,.2f}")
                    cols[4].markdown(f"📝 {ingreso['observacion'] or '—'}")
                    if cols[5].button("✏️", key=f"editar_{id_actual}"):
                        st.session_state[f"edit_{id_actual}"] = True
                        st.rerun()
                    if cols[6].button("🗑️", key=f"eliminar_{id_actual}"):
                        eliminar_ingreso(id_actual)
                        st.success("✅ Ingreso eliminado")
                        st.rerun()
    else:
        st.info("No hay ingresos registrados.")


# -------------------- PESTAÑA: REGISTRO DE GASTOS --------------------
elif menu == "💸 Registro de Gastos":
    from db_gastos import insertar_gasto, obtener_gastos, eliminar_gasto, actualizar_gasto

    st.title("💸 Registro de Gastos")
    st.subheader("Registrar nuevo gasto")

    with st.form("form_nuevo_gasto"):
        nueva_fecha = st.date_input("Fecha del gasto")
        nuevo_concepto = st.text_input("Concepto del gasto")
        nuevo_monto = st.number_input("Monto (₡)", min_value=0.0, step=1000.0, format="%.2f")
        nueva_observacion = st.text_area("Observación (opcional)")
        enviar = st.form_submit_button("Registrar")

        if enviar:
            if not nuevo_concepto.strip():
                st.warning("⚠️ El concepto no puede estar vacío.")
            elif nuevo_monto == 0:
                st.warning("⚠️ El monto no puede ser cero.")
            else:
                resultado = insertar_gasto(nueva_fecha, nuevo_concepto, nuevo_monto, nueva_observacion)
                if resultado.data:
                    st.success("✅ Gasto registrado exitosamente")
                    st.rerun()
                else:
                    st.error(f"❌ Error al registrar: {resultado.error}")

    st.subheader("📋 Gastos registrados")
    gastos = obtener_gastos()

    if gastos:
        df = pd.DataFrame(gastos)
        df["fecha"] = pd.to_datetime(df["fecha"]).dt.strftime("%d/%m/%Y")
        df["monto"] = df["monto"].map(lambda x: round(x, 2))

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Gastos")

        st.download_button(
            label="⬇️ Descargar respaldo en Excel",
            data=output.getvalue(),
            file_name="respaldo_gastos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        for gasto in gastos:
            with st.container():
                id_actual = gasto['id']
                editando = st.session_state.get(f"edit_gasto_{id_actual}", False)

                if editando:
                    st.markdown(f"### ✏️ Editando gasto ID {id_actual}")
                    fecha = st.date_input("Fecha", value=pd.to_datetime(gasto["fecha"]), key=f"fecha_gasto_{id_actual}")
                    concepto = st.text_input("Concepto", value=gasto["concepto"], key=f"concepto_gasto_{id_actual}")
                    monto = st.number_input("Monto (₡)", min_value=0.0, value=float(gasto["monto"]),
                                            format="%.2f", key=f"monto_gasto_{id_actual}")
                    observacion = st.text_input("Observación", value=gasto["observacion"], key=f"obs_gasto_{id_actual}")
                    col1, col2 = st.columns([1, 1])
                    if col1.button("💾 Guardar", key=f"guardar_gasto_{id_actual}"):
                        if not concepto.strip():
                            st.warning("⚠️ El concepto no puede estar vacío.")
                        elif monto == 0:
                            st.warning("⚠️ El monto no puede ser cero.")
                        else:
                            actualizar_gasto(id_actual, fecha, concepto, monto, observacion)
                            st.session_state[f"edit_gasto_{id_actual}"] = False
                            st.success("✅ Gasto actualizado")
                            st.rerun()
                    if col2.button("❌ Cancelar", key=f"cancelar_gasto_{id_actual}"):
                        st.session_state[f"edit_gasto_{id_actual}"] = False
                        st.rerun()

                else:
                    fecha_formateada = pd.to_datetime(gasto['fecha']).strftime("%d/%m/%Y")
                    cols = st.columns([1, 2, 2, 2, 3, 1, 1])
                    cols[0].markdown(f"**ID:** {gasto['id']}")
                    cols[1].markdown(f"📅 {fecha_formateada}")
                    cols[2].markdown(f"📄 {gasto['concepto']}")
                    cols[3].markdown(f"💰 ₡{gasto['monto']:,.2f}")
                    cols[4].markdown(f"📝 {gasto['observacion'] or '—'}")
                    if cols[5].button("✏️", key=f"editar_gasto_{id_actual}"):
                        st.session_state[f"edit_gasto_{id_actual}"] = True
                        st.rerun()
                    if cols[6].button("🗑️", key=f"eliminar_gasto_{id_actual}"):
                        eliminar_gasto(id_actual)
                        st.success("✅ Gasto eliminado")
                        st.rerun()
    else:
        st.info("No hay gastos registrados.")



# -------------------- PESTAÑA: BALANCE GENERAL --------------------
elif menu == "📊 Reporte General":
    st.title("📊 REPORTE GENERAL")
    st.write("Resumen financiero entre ingresos y gastos registrados, con filtro por rango de fechas.")

    col1, col2 = st.columns(2)
    with col1:
        fecha_inicio = st.date_input("📅 Fecha de inicio", pd.to_datetime("2025-01-01").date())
    with col2:
        fecha_fin = st.date_input("📅 Fecha de fin", pd.to_datetime("today").date())

    try:
        ingresos = obtener_ingresos()
        gastos = obtener_gastos()

        df_ingresos = pd.DataFrame(ingresos)
        df_gastos = pd.DataFrame(gastos)

        if not df_ingresos.empty:
            df_ingresos["fecha"] = pd.to_datetime(df_ingresos["fecha"]).dt.date
            df_ingresos = df_ingresos[(df_ingresos["fecha"] >= fecha_inicio) & (df_ingresos["fecha"] <= fecha_fin)]
            df_ingresos["fecha"] = df_ingresos["fecha"].apply(lambda x: x.strftime("%d/%m/%Y"))
            df_ingresos["monto"] = df_ingresos["monto"].apply(lambda x: f"₡{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            df_ingresos["observacion"] = df_ingresos["observacion"].fillna("")

        if not df_gastos.empty:
            df_gastos["fecha"] = pd.to_datetime(df_gastos["fecha"]).dt.date
            df_gastos = df_gastos[(df_gastos["fecha"] >= fecha_inicio) & (df_gastos["fecha"] <= fecha_fin)]
            df_gastos["fecha"] = df_gastos["fecha"].apply(lambda x: x.strftime("%d/%m/%Y"))
            df_gastos["monto"] = df_gastos["monto"].apply(lambda x: f"₡{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            df_gastos["observacion"] = df_gastos["observacion"].fillna("")

        st.subheader("💰 Ingresos en el período")
        st.dataframe(df_ingresos, use_container_width=True)

        st.subheader("💸 Gastos en el período")
        st.dataframe(df_gastos, use_container_width=True)

        # Cálculo del resumen
        total_ingresos = sum([i["monto"] for i in ingresos if fecha_inicio <= pd.to_datetime(i["fecha"]).date() <= fecha_fin])
        total_gastos = sum([i["monto"] for i in gastos if fecha_inicio <= pd.to_datetime(i["fecha"]).date() <= fecha_fin])
        balance_final = total_ingresos - total_gastos
        color = "green" if balance_final >= 0 else "red"

        st.markdown("---")
        st.markdown("### 🪙 Resumen del período seleccionado:")
        st.markdown(f"**Total de ingresos:** ₡{total_ingresos:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        st.markdown(f"**Total de gastos:** ₡{total_gastos:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        st.markdown(
            f"<strong>Balance final:</strong> <span style='color:{color}; font-weight:bold;'>₡{balance_final:,.2f}</span>"
            .replace(",", "X").replace(".", ",").replace("X", "."),
            unsafe_allow_html=True
        )

    except Exception as e:
        st.error("⚠️ Error al obtener o procesar los datos. Verifica la conexión con Supabase o el formato de datos.")


# -------------------- PESTAÑA: Generador de PDF --------------------
elif menu == "📄 Exportar PDF":
    from fpdf import FPDF
    import datetime

    st.title("📄 Exportar PDF del Informe Financiero")
    st.write("Genera un PDF con el resumen de ingresos y gastos para un período seleccionado.")

    col1, col2 = st.columns(2)
    fecha_inicio = col1.date_input("📅 Fecha de inicio", value=datetime.date(2025, 1, 1))
    fecha_fin = col2.date_input("📅 Fecha de fin", value=datetime.date(2025, 6, 30))

    if st.button("📥 Generar PDF"):
        try:
            ingresos = obtener_ingresos()
            gastos = obtener_gastos()

            ingresos_filtrados = [i for i in ingresos if fecha_inicio <= datetime.datetime.strptime(i["fecha"], "%Y-%m-%d").date() <= fecha_fin]
            gastos_filtrados = [g for g in gastos if fecha_inicio <= datetime.datetime.strptime(g["fecha"], "%Y-%m-%d").date() <= fecha_fin]

            total_ingresos = sum(i.get("monto", 0.0) or 0.0 for i in ingresos_filtrados)
            total_gastos = sum(g.get("monto", 0.0) or 0.0 for g in gastos_filtrados)
            balance = total_ingresos - total_gastos

            class PDF(FPDF):
                def header(self):
                    self.set_fill_color(0, 102, 204)
                    self.set_text_color(255, 255, 255)
                    self.set_font("Arial", "B", 16)
                    self.cell(0, 12, "Informe Financiero", ln=True, align="C", fill=True)

                def footer(self):
                    self.set_y(-15)
                    self.set_font("Arial", "I", 8)
                    self.set_text_color(128)
                    self.cell(0, 10, f"Página {self.page_no()} - Iglesia Restauración Colonia Carvajal", 0, 0, "C")

            pdf = PDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.set_text_color(0)
            pdf.ln(5)

            # Texto institucional
            pdf.set_font("Arial", "", 11)
            pdf.multi_cell(0, 10, "Este informe fue solicitado por los pastores Jeannett Loaiciga Segura y Carlos Castro Campos.", align="L")
            pdf.ln(3)
            pdf.set_font("Arial", "I", 10)
            pdf.cell(0, 10, f"Período: {fecha_inicio.strftime('%d-%m-%Y')} al {fecha_fin.strftime('%d-%m-%Y')}.", ln=True)

            # Ingresos
            pdf.ln(6)
            pdf.set_fill_color(204, 229, 255)
            pdf.set_font("Arial", "B", 13)
            pdf.cell(0, 10, "Ingresos", ln=True, fill=True)

            pdf.set_font("Arial", "", 11)
            pdf.set_draw_color(0, 102, 204)
            pdf.set_line_width(0.5)
            pdf.ln(2)
            if ingresos_filtrados:
                for i in ingresos_filtrados:
                    fecha = datetime.datetime.strptime(i.get("fecha", ""), "%Y-%m-%d").strftime("%d-%m-%Y")
                    tipo = i.get("concepto") or "Sin tipo"
                    monto = i.get("monto") or 0.0
                    detalle = i.get("observacion") or "Sin detalle"
                    pdf.multi_cell(0, 8, f"{fecha}: {tipo} - CRC {monto:,.2f} - {detalle}", border=1)
            else:
                pdf.cell(0, 10, "No se registraron ingresos en este período.", ln=True)

            pdf.set_font("Arial", "B", 11)
            pdf.cell(0, 10, f"Total ingresos: CRC {total_ingresos:,.2f}.", ln=True)

            # Gastos
            pdf.ln(6)
            pdf.set_fill_color(255, 204, 204)
            pdf.set_font("Arial", "B", 13)
            pdf.cell(0, 10, "Gastos", ln=True, fill=True)

            pdf.set_font("Arial", "", 11)
            pdf.set_draw_color(204, 0, 0)
            pdf.set_line_width(0.5)
            pdf.ln(2)
            if gastos_filtrados:
                for g in gastos_filtrados:
                    fecha = datetime.datetime.strptime(g.get("fecha", ""), "%Y-%m-%d").strftime("%d-%m-%Y")
                    tipo = g.get("concepto") or "Sin tipo"
                    monto = g.get("monto") or 0.0
                    detalle = g.get("observacion") or "Sin detalle"
                    pdf.multi_cell(0, 8, f"{fecha}: {tipo} - CRC {monto:,.2f} - {detalle}", border=1)
            else:
                pdf.cell(0, 10, "No se registraron gastos en este período.", ln=True)

            pdf.set_font("Arial", "B", 11)
            pdf.cell(0, 10, f"Total gastos: CRC {total_gastos:,.2f}.", ln=True)

            # Balance final
            pdf.ln(10)
            pdf.set_fill_color(224, 255, 224)
            pdf.set_font("Arial", "B", 13)
            pdf.cell(0, 10, f"Balance final: CRC {balance:,.2f}.", ln=True, fill=True)

            # Firmas
            pdf.ln(20)
            pdf.set_font("Arial", "", 11)
            pdf.cell(80, 10, "Firma Pastora Jeannett Loaiciga Segura", ln=0, align="C")
            pdf.cell(30, 10, "", ln=0)
            pdf.cell(80, 10, "Firma Pastor Carlos Castro Campos", ln=1, align="C")
            pdf.cell(80, 10, "______________________________", ln=0, align="C")
            pdf.cell(30, 10, "", ln=0)
            pdf.cell(80, 10, "______________________________", ln=1, align="C")

            # Descargar PDF
            pdf_output = pdf.output(dest="S").encode("latin-1", "ignore")
            st.download_button("📩 Descargar PDF", data=pdf_output, file_name="informe_financiero.pdf", mime="application/pdf")

        except Exception as e:
            st.error(f"❌ Error al generar el PDF: {e}")


















